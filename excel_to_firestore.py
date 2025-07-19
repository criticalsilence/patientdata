import openpyxl # openpyxl'i içe aktarın
import firebase_admin
from firebase_admin import credentials, firestore
import os

# --- YAPILANDIRMA AYARLARI ---
SERVICE_ACCOUNT_KEY_PATH = 'serviceAccountKey.json'
EXCEL_FILE_PATH = 'hasta_datalari.xlsx'
FIRESTORE_COLLECTION_NAME = 'hasta_bilgileri'
# -----------------------------

def initialize_firebase():
    """Firebase Admin SDK'yı başlatır."""
    if not os.path.exists(SERVICE_ACCOUNT_KEY_PATH):
        print(f"Hata: Firebase servis hesabı anahtar dosyası bulunamadı: {SERVICE_ACCOUNT_KEY_PATH}")
        print("Lütfen Firebase konsolundan 'serviceAccountKey.json' dosyasını indirin ve doğru yolu belirtin.")
        return None

    try:
        cred = credentials.Certificate(SERVICE_ACCOUNT_KEY_PATH)
        firebase_admin.initialize_app(cred)
        db = firestore.client()
        print("Firebase başarıyla başlatıldı.")
        return db
    except Exception as e:
        print(f"Firebase başlatılırken bir hata oluştu: {e}")
        return None

def read_excel_data_without_pandas(file_path, header_row_index=1):
    """
    Excel dosyasını openpyxl kullanarak okur ve bir liste sözlük döndürür.
    header_row_index: Başlıkların bulunduğu satırın 0-tabanlı indeksi.
                      (Excel'de 2. satır ise, burası 1 olmalı)
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        # Başlıkları oku
        headers = [cell.value for cell in sheet[header_row_index + 1]] # Excel'de 2. satır (index 1) ise, +1 ekleriz.
        
        data_rows = []
        # Veri satırlarını oku (başlık satırından sonra başla)
        for row_index in range(header_row_index + 2, sheet.max_row + 1): # Başlık satırından sonraki satırdan başla
            row_data = {}
            for col_index, header in enumerate(headers):
                cell_value = sheet.cell(row=row_index, column=col_index + 1).value
                # Boş hücreleri None olarak kaydet
                row_data[header] = cell_value if cell_value is not None else None
            data_rows.append(row_data)

        print(f"'{file_path}' Excel dosyası openpyxl ile başarıyla okundu. Toplam {len(data_rows)} satır.")
        return data_rows
    except FileNotFoundError:
        print(f"Hata: Excel dosyası bulunamadı: {file_path}")
        return None
    except Exception as e:
        print(f"Excel dosyası okunurken bir hata oluştu: {e}")
        return None

def upload_to_firestore(db_client, data_list, collection_name):
    """Liste sözlükleri Firestore'a yükler."""
    print(f"Veriler '{collection_name}' koleksiyonuna aktarılıyor...")
    success_count = 0
    fail_count = 0

    for index, data in enumerate(data_list):
        try:
            doc_ref = db_client.collection(collection_name).add(data)
            print(f"Satır {index+1} aktarıldı. Belge ID: {doc_ref[1].id}")
            success_count += 1
        except Exception as e:
            print(f"Hata: Satır {index+1} aktarılırken bir hata oluştu: {e} - Veri: {data}")
            fail_count += 1

    print("\n--- Veri Aktarımı Tamamlandı ---")
    print(f"Başarıyla aktarılan belge sayısı: {success_count}")
    print(f"Hata oluşan belge sayısı: {fail_count}")
    print(f"Firestore'daki '{collection_name}' koleksiyonunuzu kontrol edebilirsiniz.")

if __name__ == "__main__":
    db = initialize_firebase()
    if db:
        # Başlıkların 2. satırda (0-tabanlı indeks 1) olduğunu varsayıyoruz.
        # Eğer Excel'de başlıklar 1. satırda ise header_row_index=0 yapın.
        data_to_upload = read_excel_data_without_pandas(EXCEL_FILE_PATH, header_row_index=1)
        if data_to_upload is not None:
            upload_to_firestore(db, data_to_upload, FIRESTORE_COLLECTION_NAME)

